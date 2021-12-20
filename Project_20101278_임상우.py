from openpyxl.drawing.image import Image
import re
import os
import openpyxl
import urllib.request as req
from bs4 import BeautifulSoup



if not os.path.exists("./MelonImage"):
    os.mkdir("./MelonImage")

head = req.Request("https://www.melon.com/chart/index.htm", headers={"User-Agent":"Mozilla/5.0"})
url = req.urlopen(head)
soupcode = BeautifulSoup(url, "html.parser")
title = soupcode.select("div.ellipsis.rank01 > span > a")
img = soupcode.select("a.image_typeAll > img")
name = soupcode.select("div.ellipsis.rank02 > span")
album = soupcode.select("div.ellipsis.rank03 > a")


if not os.path.exists("./MelonRanking.xlsx"):
    openpyxl.Workbook().save("./MelonRanking.xlsx")

page = openpyxl.load_workbook("./MelonRanking.xlsx")

if "Sheet" in page.sheetnames:
    page.remove(page["Sheet"])
p1 = page.create_sheet()
p1.title = "멜론 음원사이트 차트 순위표"
row_num = 1

p1.column_dimensions["A"].width = 15
p1.column_dimensions["B"].width = 50
p1.column_dimensions["C"].width = 29
p1.column_dimensions["D"].width = 47

for i in range(len(title)):
    img_dir = "./MelonImage"
    img_file_name = img_dir + "/" + re.sub("[\\\/:*?\"<>\|]", " ", title[i].string) + ".png"
    req.urlretrieve(img[i].attrs["src"], img_file_name)

    img_for_excel = Image(img_file_name)
    p1.add_image(img_for_excel, "A{}".format(row_num))
    p1.cell(row=row_num, column=2).value = title[i].string
    p1.cell(row=row_num, column=3).value = name[i].text
    p1.cell(row=row_num, column=4).value = album[i].string
    p1.row_dimensions[row_num].height = 90
    page.save("./MelonRanking.xlsx")
    print("{}위. {} - {}".format(row_num, title[i].string, name[i].text))
    row_num += 1
